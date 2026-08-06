"""
Ingest complete STME Odd Semester 2026-27 timetable into Supabase.

Populates:
  subjects.code           — short code (CAL, DM, …)
  faculty_members         — 43 faculty with code + total hours
  venues                  — 11 classrooms + 13 labs
  timetable_slots         — every scheduled session (day/time/subject/faculty/room)

Run from the project root:
    python3 scripts/ingest_full_timetable.py
"""

import re
import uuid
import datetime
import psycopg2
import openpyxl
from pathlib import Path

# ─── Config ────────────────────────────────────────────────────────────────────
DB_URL = "postgresql://postgres:8591282986110418@db.nnbveokfgsbcnrnzziny.supabase.co:5432/postgres"
EXCEL_PATH = Path(__file__).parent.parent / "STME_Odd_Sem_2026-27 (3).xlsm"
ACADEMIC_YEAR = "2026-27"

# Sheet → (programme_code, semester, division)
SHEET_MAP = {
    "MBA TECH CE SEM 1":  ("MBA_TECH",   1, None),
    "B TECH CE-A SEM 1":  ("BTECH_CE",   1, "A"),
    "B TECH CE-B SEM 1":  ("BTECH_CE",   1, "B"),
    "B TECH AI&DS SEM 1": ("BTECH_AIDS", 1, None),
    "MBA TECH CE SEM 3":  ("MBA_TECH",   3, None),
    "B TECH CE-A SEM 3":  ("BTECH_CE",   3, "A"),
    "B TECH CE-B SEM 3":  ("BTECH_CE",   3, "B"),
    "B TECH AI&DS SEM 3": ("BTECH_AIDS", 3, None),
    "MBA TECH CE SEM 5":  ("MBA_TECH",   5, None),
    "B TECH CE-A SEM 5":  ("BTECH_CE",   5, "A"),
    "B TECH CE-B SEM 5":  ("BTECH_CE",   5, "B"),
    "B TECH AI&DS SEM 5": ("BTECH_AIDS", 5, None),
    "MBA TECH CE SEM 7":  ("MBA_TECH",   7, None),
    "B TECH CE SEM 7":    ("BTECH_CE",   7, None),
    "B TECH AI&DS SEM 7": ("BTECH_AIDS", 7, None),
    "MBA TECH CE SEM 9":  ("MBA_TECH",   9, None),
}

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Known visiting faculty codes
VISITING_CODES = {
    "PPVF", "QJ", "RZ", "RPT", "BB", "TSH", "AH", "CP", "AN",
    "SSF", "AS", "CM", "RK", "RP", "SAG", "MN", "PDR", "KP", "IAK",
}

# ─── Subject code extraction ────────────────────────────────────────────────────

def extract_subject_code(raw_name: str) -> str | None:
    """Extract the short code from 'Full Name (CODE)' or 'PREFIX: Full Name (CODE)'."""
    raw = str(raw_name).strip()
    # strip elective prefix
    raw = re.sub(r"^[A-Z]{1,3}\d?: ", "", raw)
    # extract trailing (CODE)
    m = re.search(r"\(([A-Z][A-Z0-9&/]{0,12})\)\s*$", raw)
    return m.group(1) if m else None


def extract_code_from_subject_table(ws):
    """Return {cleaned_name: code} from the programme sheet subject table."""
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    in_table = False
    for row in rows:
        if row[0] and "Subject Name" in str(row[0]):
            in_table = True
            continue
        if not in_table:
            continue
        if row[0] and (str(row[0]).startswith("Total") or str(row[0]).startswith("Class")):
            break
        if not row[0] or str(row[0]).strip() in ("", "None", "Theory"):
            continue
        raw = str(row[0]).strip()
        code = extract_subject_code(raw)
        if code:
            # clean name (same logic as ingest_timetable.py)
            name = re.sub(r"^[A-Z]{1,3}\d?: ", "", raw)
            name = re.sub(r"\s*\([A-Z][A-Z0-9&/\s]{0,12}\)\s*$", "", name)
            name = name.replace("Enviromental", "Environmental").strip()
            result[name] = code
    return result


# ─── Faculty extraction ─────────────────────────────────────────────────────────

def extract_faculty_members(wb) -> list[dict]:
    """Extract one record per faculty sheet."""
    all_sheets = wb.sheetnames
    skip = set(SHEET_MAP.keys()) | {
        "T5", "C101", "C102", "C106", "C107", "C307", "C308", "C309",
        "C310", "C311", "C312", "C316",
        "L101", "L102", "L103", "L104", "L106", "L108", "L109", "L110",
        "L203", "L204", "L205", "L209", "L210",
        "_Template_Lab", "_Template_Classroom", "_Template_Faculty", "_TimeGrid_ID",
    }
    faculty = []
    for sheet_name in all_sheets:
        if sheet_name in skip:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        full_name = None
        total_hours = None
        for row in rows:
            if row[0] and "Faculty Name:" in str(row[0]):
                full_name = str(row[0]).replace("Faculty Name:", "").strip().rstrip("\t ")
            if row[0] and "Total hours" in str(row[0]):
                vals = [v for v in row if v is not None]
                total_hours = int(vals[-1]) if vals and isinstance(vals[-1], (int, float)) else None
        if full_name:
            code = sheet_name
            ftype = "Visiting" if code in VISITING_CODES else "Core"
            faculty.append({
                "code": code,
                "full_name": full_name,
                "faculty_type": ftype,
                "weekly_hours_total": total_hours,
            })
    return faculty


# ─── Venue extraction ───────────────────────────────────────────────────────────

def build_venues() -> list[dict]:
    classrooms = ["C101", "C102", "C106", "C107", "C307", "C308",
                  "C309", "C310", "C311", "C312", "C316", "T5"]
    labs = ["L101", "L102", "L103", "L104", "L106", "L108", "L109",
            "L110", "L203", "L204", "L205", "L209", "L210"]
    venues = []
    for code in classrooms:
        # T5 → T-5, C101 → C-101
        if code.startswith("T"):
            label = f"T-{code[1:]}"
        else:
            label = f"C-{code[1:]}"
        venues.append({"code": code, "label": label, "venue_type": "classroom"})
    for code in labs:
        label = f"L-{code[1:]}"
        venues.append({"code": code, "label": label, "venue_type": "lab"})
    return venues


# ─── Timetable slot parsing ─────────────────────────────────────────────────────

def parse_slot_cell(cell_text: str) -> dict | None:
    """Parse 'SUBJ\n(FAC)\nVENUE' or 'SUBJ (T) B1\n(FAC)\nVENUE' etc."""
    lines = [ln.strip() for ln in str(cell_text).split("\n") if ln.strip()]
    if not lines:
        return None

    first = lines[0]

    # slot_type from '(T)' in first line
    is_tutorial = "(T)" in first
    first_clean = first.replace("(T)", "").strip()

    # strip elective prefix (DE4: , OE1: , PE1: etc.)
    first_clean = re.sub(r"^[A-Z]{1,3}\d?: ", "", first_clean).strip()

    # batch from first line e.g. "B1", "B2", "B3"
    batch_m = re.search(r"\bB([123])\b", first_clean)
    batch = f"B{batch_m.group(1)}" if batch_m else None
    if batch:
        first_clean = first_clean.replace(batch, "").strip()

    # subject code is the leading uppercase token
    subj_m = re.match(r"^([A-Z][A-Z0-9&/]*)", first_clean)
    subject_code = subj_m.group(1) if subj_m else first_clean

    # remaining lines: (FACULTY_CODE) and VENUE
    faculty_code = None
    venue_label = None
    for line in lines[1:]:
        if re.match(r"^\([A-Z]+\)$", line):
            faculty_code = line[1:-1]
        elif re.match(r"^[BCLT]-?\d+$", line):
            venue_label = line
        elif re.match(r"^B[123]$", line):
            batch = line

    # infer slot_type
    if is_tutorial:
        slot_type = "Tutorial"
    elif venue_label and re.match(r"^L-?\d+", venue_label):
        slot_type = "Practical"
    else:
        slot_type = "Lecture"

    return {
        "subject_code": subject_code,
        "faculty_code": faculty_code,
        "venue_label": venue_label,
        "batch": batch,
        "slot_type": slot_type,
    }


def extract_slots_from_sheet(ws, prog_code: str, sem: int, division: str | None) -> list[dict]:
    """Extract all timetable slots from a programme sheet."""
    rows = list(ws.iter_rows(values_only=True))

    # Find day-header row
    day_header_idx = None
    for i, row in enumerate(rows):
        if row[0] and "Time/Day" in str(row[0]):
            day_header_idx = i
            break
    if day_header_idx is None:
        return []

    day_header = rows[day_header_idx]

    # Build column → day name mapping
    col_to_day = {}
    current_day = None
    for col, val in enumerate(day_header):
        if val and str(val).strip() in DAYS:
            current_day = str(val).strip()
        if current_day and col >= 2:
            col_to_day[col] = current_day

    slots = []
    for row in rows[day_header_idx + 1:]:
        time_start = row[0]
        time_end   = row[1] if len(row) > 1 else None

        # Only process rows with actual time values
        if not isinstance(time_start, datetime.time):
            continue
        # Skip break rows
        if len(row) > 2 and isinstance(row[2], str) and "Break" in row[2]:
            continue

        for col, cell_val in enumerate(row):
            if col < 2 or col not in col_to_day:
                continue
            if not cell_val or not isinstance(cell_val, str):
                continue

            day = col_to_day[col]
            parsed = parse_slot_cell(cell_val)
            if not parsed or not parsed["subject_code"]:
                continue

            slots.append({
                "prog_code":    prog_code,
                "sem":          sem,
                "division":     division,
                "day_of_week":  day,
                "slot_start":   time_start,
                "slot_end":     time_end,
                **parsed,
            })

    return slots


# ─── Main ingestion ─────────────────────────────────────────────────────────────

def main():
    print(f"Loading workbook …")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), keep_vba=True, data_only=True)

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # ── Load lookup maps ────────────────────────────────────────────────────────
    cur.execute("SELECT code, id FROM programmes")
    prog_map = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT id, programme_id, name, semester_number FROM subjects")
    subjects_db = cur.fetchall()  # (id, prog_id, name, sem)

    # ── 1. Update subjects.code ─────────────────────────────────────────────────
    print("\n[1/4] Extracting subject codes …")
    # Build name→code map from each programme sheet
    name_to_code: dict[str, str] = {}
    for sheet_name in SHEET_MAP:
        ws = wb[sheet_name]
        name_to_code.update(extract_code_from_subject_table(ws))

    updated = 0
    for subj_id, prog_id, name, sem in subjects_db:
        code = name_to_code.get(name)
        if code:
            cur.execute("UPDATE subjects SET code=%s WHERE id=%s", (code, subj_id))
            updated += 1
    conn.commit()
    print(f"  Updated {updated} subjects with codes")

    # Build (prog_id, sem, subject_code) → subject_id lookup
    cur.execute("SELECT id, programme_id, semester_number, code FROM subjects WHERE code IS NOT NULL")
    subj_lookup: dict[tuple, str] = {}
    for sid, pid, sem, code in cur.fetchall():
        subj_lookup[(str(pid), sem, code)] = str(sid)

    # ── 2. Faculty members ──────────────────────────────────────────────────────
    print("\n[2/4] Inserting faculty_members …")
    cur.execute("TRUNCATE TABLE faculty_members CASCADE")
    conn.commit()

    faculty_list = extract_faculty_members(wb)
    faculty_id_map: dict[str, str] = {}  # code → uuid
    for f in faculty_list:
        fid = str(uuid.uuid4())
        cur.execute(
            """
            INSERT INTO faculty_members (id, code, full_name, faculty_type, weekly_hours_total)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (fid, f["code"], f["full_name"], f["faculty_type"], f["weekly_hours_total"]),
        )
        faculty_id_map[f["code"]] = fid
    conn.commit()
    print(f"  Inserted {len(faculty_list)} faculty members")

    # ── 3. Venues ───────────────────────────────────────────────────────────────
    print("\n[3/4] Inserting venues …")
    cur.execute("TRUNCATE TABLE venues CASCADE")
    conn.commit()

    venue_list = build_venues()
    venue_id_map: dict[str, str] = {}  # label → uuid  (e.g. 'C-101' → uuid)
    for v in venue_list:
        vid = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO venues (id, code, label, venue_type) VALUES (%s, %s, %s, %s)",
            (vid, v["code"], v["label"], v["venue_type"]),
        )
        venue_id_map[v["label"]] = vid
    conn.commit()
    print(f"  Inserted {len(venue_list)} venues")

    # ── 4. Timetable slots ──────────────────────────────────────────────────────
    print("\n[4/4] Extracting & inserting timetable_slots …")
    cur.execute("TRUNCATE TABLE timetable_slots CASCADE")
    conn.commit()

    all_slots = []
    for sheet_name, (prog_code, sem, division) in SHEET_MAP.items():
        ws = wb[sheet_name]
        slots = extract_slots_from_sheet(ws, prog_code, sem, division)
        all_slots.extend(slots)

    print(f"  Raw slots extracted: {len(all_slots)}")

    inserted = 0
    skipped = 0
    for slot in all_slots:
        prog_id  = prog_map.get(slot["prog_code"])
        if not prog_id:
            skipped += 1
            continue

        # Resolve subject_id
        subj_id = subj_lookup.get((str(prog_id), slot["sem"], slot["subject_code"]))

        # Resolve faculty_id
        fac_id = faculty_id_map.get(slot["faculty_code"]) if slot["faculty_code"] else None

        # Resolve venue_id
        venue_id = venue_id_map.get(slot["venue_label"]) if slot["venue_label"] else None

        sid = str(uuid.uuid4())
        cur.execute(
            """
            INSERT INTO timetable_slots (
                id, academic_year, programme_id, semester_number, division,
                day_of_week, slot_start, slot_end,
                subject_code, subject_id, faculty_code, faculty_id,
                venue_label, venue_id, batch, slot_type
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s
            )
            """,
            (
                sid, ACADEMIC_YEAR, prog_id, slot["sem"], slot["division"],
                slot["day_of_week"], slot["slot_start"], slot["slot_end"],
                slot["subject_code"], subj_id, slot["faculty_code"], fac_id,
                slot["venue_label"], venue_id, slot["batch"], slot["slot_type"],
            ),
        )
        inserted += 1

    conn.commit()
    print(f"  Inserted {inserted} slots  |  Skipped {skipped}")

    # ── Summary ─────────────────────────────────────────────────────────────────
    print("\n=== Final row counts ===")
    for t in ["programmes", "subjects", "faculty_members", "venues",
              "faculty_workload", "timetable_slots"]:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        print(f"  {t:25s}: {cur.fetchone()[0]}")

    cur.close()
    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
